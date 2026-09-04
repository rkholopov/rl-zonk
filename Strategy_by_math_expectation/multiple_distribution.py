from copy import deepcopy
from openpyxl import load_workbook
import pandas as pd

dist = pd.read_excel('Distribution_of_SBME.xlsx', sheet_name='Sheet1')
dist = dist[0].tolist()

compact_dist = [[int(dist[0]) - int(dist[0]) % 50, 1 / 6**6]]
for i in range(1, len(dist)):
    if int(dist[i]) - int(dist[i]) % 50 != compact_dist[-1][0]:
        compact_dist.append([int(dist[i]) - int(dist[i]) % 50, 1 / 6**6])
    else:
        compact_dist[-1][1] += 1 / 6**6

mult_dist = deepcopy(compact_dist)
for k in range(9):
    next_mult_dist = []
    dict_next_mult_dist = dict()
    for i in range(len(mult_dist)):
        for j in range(len(compact_dist)):
            x = int(mult_dist[i][0] + compact_dist[j][0]) - int(mult_dist[i][0] + compact_dist[j][0]) % 50
            cnt = mult_dist[i][1] * compact_dist[j][1]
            if x in dict_next_mult_dist.keys():
                next_mult_dist[dict_next_mult_dist[x]][1] += cnt
            else:
                dict_next_mult_dist[x] = len(next_mult_dist)
                next_mult_dist.append([x, cnt])

    mult_dist = deepcopy(next_mult_dist)

'''
wb = load_workbook("Distribution_of_SBME.xlsx")
ws = wb.create_sheet("dist_10_rounds")
for row in mult_dist:
    ws.append(row)

wb.save("Distribution_of_SBME.xlsx")
'''
